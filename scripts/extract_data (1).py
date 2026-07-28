"""
extract_data.py
----------------
แปลงไฟล์ 1.xlsx - 7.xlsx (ผลสำรวจเห็ดเขาหลวง) ให้เป็น:
  - data/round-1.json ... round-7.json
  - data/summary.json
  - public/images/round-1/ ... round-7/  (รูปเห็ดที่ดึงออกจากไฟล์ excel)

วิธีรัน:
    pip install openpyxl --break-system-packages
    python3 scripts/extract_data.py

ค่าเริ่มต้น: อ่านไฟล์จากโฟลเดอร์ source-data/1.xlsx ... 7.xlsx
ถ้าไฟล์ของคุณอยู่ที่อื่น แก้ตัวแปร SOURCE_DIR ด้านล่าง
"""

import json
import os
import re
import statistics
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ---------- ตั้งค่า path ----------
SOURCE_DIR = "source-data"       # โฟลเดอร์ที่เก็บไฟล์ 1.xlsx - 7.xlsx
DATA_OUT_DIR = "data"
IMAGES_OUT_DIR = os.path.join("public", "images")

THAI_DATES = {
    1: "22 พฤศจิกายน 2568",
    2: "29 ธันวาคม 2568",
    3: "16 มกราคม 2569",
    4: "14 กุมภาพันธ์ 2569",
    5: "7 พฤษภาคม 2569",
    6: "18 มิถุนายน 2569",
    7: "18 กรกฎาคม 2569",
}

# คอลัมน์ (index เริ่มที่ 0) ตามโครงสร้างไฟล์จริง
COL_POINT = 0          # A จุดที่สำรวจ
COL_IMAGE = 1           # B ภาพเห็ด (เป็นรูปฝัง ไม่ใช่ text)
COL_FAMILY = 2          # C วงศ์เห็ด
COL_SCI_NAME = 3        # D ชื่อวิทยาศาสตร์
COL_LOCAL_NAME = 4      # E ชื่อท้องถิ่น
COL_AMOUNT = 5          # F ปริมาณที่พบ
COL_GROUP = 6           # G กลุ่มเห็ด
COL_ORIGIN = 7          # H แหล่งกำเนิดเห็ด
COL_ROLE = 8            # I หน้าที่ในระบบนิเวศ
COL_EDIBILITY = 9       # J การรับประทาน
COL_TEMP = 10           # K อุณหภูมิอากาศ
COL_HUMIDITY = 11       # L ความชื้นสัมพัทธ์อากาศ


def normalize_name(name):
    """ตัดช่องว่างหัวท้าย/ซ้ำ เพื่อใช้เทียบชื่อวิทยาศาสตร์"""
    if not name:
        return ""
    return re.sub(r"\s+", " ", str(name).strip())


def find_data_end_row(ws):
    """หาว่าตารางข้อมูลดิบสิ้นสุดที่แถวไหน (ก่อนถึงบล็อกสรุป 'วิเคราะห์ข้อมูลในภาพรวม')"""
    for row_idx in range(4, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if val and "วิเคราะห์ข้อมูลในภาพรวม" in str(val):
            return row_idx - 1
    return ws.max_row


def extract_images_by_row(ws):
    """คืน dict {row_number: [image_object, ...]} จากรูปที่ฝังในชีต"""
    images_by_row = {}
    for img in getattr(ws, "_images", []):
        try:
            anchor_row = img.anchor._from.row + 1  # openpyxl เก็บ row เริ่มที่ 0
        except AttributeError:
            continue
        images_by_row.setdefault(anchor_row, []).append(img)
    return images_by_row


def save_image(img, out_path):
    """บันทึกไฟล์รูปจาก openpyxl image object"""
    data = img._data()
    with open(out_path, "wb") as f:
        f.write(data)


def process_round(round_num, source_path, images_out_root):
    wb = load_workbook(source_path)
    ws = wb.active

    end_row = find_data_end_row(ws)
    images_by_row = extract_images_by_row(ws)

    round_img_dir = os.path.join(images_out_root, f"round-{round_num}")
    os.makedirs(round_img_dir, exist_ok=True)

    species = {}  # key = normalized scientific name -> record
    temps, humidities = [], []

    for row_idx in range(4, end_row + 1):
        point = ws.cell(row=row_idx, column=COL_POINT + 1).value
        sci_name_raw = ws.cell(row=row_idx, column=COL_SCI_NAME + 1).value
        if not point or not sci_name_raw:
            continue

        sci_key = normalize_name(sci_name_raw).lower()
        if not sci_key:
            continue

        temp = ws.cell(row=row_idx, column=COL_TEMP + 1).value
        humidity = ws.cell(row=row_idx, column=COL_HUMIDITY + 1).value
        if isinstance(temp, (int, float)):
            temps.append(temp)
        if isinstance(humidity, (int, float)):
            humidities.append(humidity)

        # ดึงรูปของแถวนี้ถ้ามี
        row_images = []
        if row_idx in images_by_row:
            for i, img in enumerate(images_by_row[row_idx]):
                filename = f"round-{round_num}-row-{row_idx:04d}-{i}.png"
                out_path = os.path.join(round_img_dir, filename)
                if not os.path.exists(out_path):
                    save_image(img, out_path)
                row_images.append(f"/images/round-{round_num}/{filename}")

        if sci_key not in species:
            species[sci_key] = {
                "scientificName": normalize_name(sci_name_raw),
                "localName": ws.cell(row=row_idx, column=COL_LOCAL_NAME + 1).value or "-",
                "family": ws.cell(row=row_idx, column=COL_FAMILY + 1).value or "-",
                "group": ws.cell(row=row_idx, column=COL_GROUP + 1).value or "-",
                "habitat": ws.cell(row=row_idx, column=COL_ORIGIN + 1).value or "-",
                "ecologicalRole": ws.cell(row=row_idx, column=COL_ROLE + 1).value or "-",
                "edibility": ws.cell(row=row_idx, column=COL_EDIBILITY + 1).value or "ไม่มีข้อมูล",
                "totalFound": 0,
                "pointsFound": [],
                "images": [],
            }

        rec = species[sci_key]
        rec["totalFound"] += 1
        if point not in rec["pointsFound"]:
            rec["pointsFound"].append(point)
        if row_images:
            rec["images"].extend(row_images)

    mushroom_list = []
    for rec in species.values():
        rec["pointsFoundCount"] = len(rec["pointsFound"])
        if not rec["images"]:
            rec["images"] = ["/images/placeholder-mushroom.png"]
        mushroom_list.append(rec)

    # เรียงตามชื่อวิทยาศาสตร์
    mushroom_list.sort(key=lambda m: m["scientificName"].lower())

    avg_temp = round(statistics.mean(temps), 1) if temps else None
    avg_humidity = round(statistics.mean(humidities), 1) if humidities else None

    result = {
        "round": round_num,
        "date": THAI_DATES.get(round_num, ""),
        "speciesCount": len(mushroom_list),
        "mushrooms": mushroom_list,
        "avgTemperature": avg_temp,
        "avgHumidity": avg_humidity,
    }
    return result


def main():
    os.makedirs(DATA_OUT_DIR, exist_ok=True)
    os.makedirs(IMAGES_OUT_DIR, exist_ok=True)

    summary = []

    for round_num in range(1, 8):
        source_path = os.path.join(SOURCE_DIR, f"{round_num}.xlsx")
        if not os.path.exists(source_path):
            print(f"⚠️  ไม่พบไฟล์ {source_path} ข้ามรอบนี้ไป")
            continue

        print(f"กำลังประมวลผลรอบที่ {round_num} ...")
        result = process_round(round_num, source_path, IMAGES_OUT_DIR)

        out_json_path = os.path.join(DATA_OUT_DIR, f"round-{round_num}.json")
        with open(out_json_path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)

        print(f"  -> พบ {result['speciesCount']} ชนิด, "
              f"เฉลี่ยอุณหภูมิ {result['avgTemperature']}°C, "
              f"เฉลี่ยความชื้น {result['avgHumidity']}%")

        summary.append({
            "round": round_num,
            "date": result["date"],
            "speciesCount": result["speciesCount"],
            "avgTemperature": result["avgTemperature"],
            "avgHumidity": result["avgHumidity"],
        })

    with open(os.path.join(DATA_OUT_DIR, "summary.json"), "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print("\nเสร็จสิ้น! ไฟล์ข้อมูลอยู่ในโฟลเดอร์ data/ และรูปอยู่ใน public/images/")


if __name__ == "__main__":
    main()

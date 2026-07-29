"""
extract_data.py (เวอร์ชัน 3 - เพิ่มข้อมูลอุณหภูมิ/ความชื้น แยกเห็ดดิน-เห็ดตอไม้ + แก้ตัวพิมพ์ใหญ่ชื่อวิทยาศาสตร์)
----------------
แปลงไฟล์ 1.xlsx - 7.xlsx (ผลสำรวจเห็ดเขาหลวง) ให้เป็น:
  - data/round-1.json ... round-7.json
  - data/summary.json
  - public/images/round-1/ ... round-7/

วิธีรัน:
    pip install openpyxl --break-system-packages
    python3 scripts/extract_data.py

สิ่งที่เพิ่มจากเวอร์ชันก่อนหน้า:
  1. ตรวจว่าเห็ดแต่ละชนิดเป็น "เห็ดดิน" หรือ "เห็ดตอไม้" จากคอลัมน์ M/N/O
     (pH ดิน, อุณหภูมิดิน, ความชื้นดิน) — ถ้าแถวไหนมีค่าคอลัมน์เหล่านี้ = เห็ดดิน
  2. คำนวณค่าเฉลี่ยอุณหภูมิ/ความชื้นอากาศ ต่อชนิดเห็ด (จากทุกแถวที่พบชนิดนั้น)
  3. คำนวณค่าเฉลี่ย pH ดิน/อุณหภูมิดิน/ความชื้นดิน ต่อชนิดเห็ด (เฉพาะชนิดที่เป็นเห็ดดิน)
  4. ทำให้ชื่อวิทยาศาสตร์ขึ้นต้นด้วยตัวพิมพ์ใหญ่เสมอ (ตัวอักษรตัวแรกเท่านั้น
     ส่วนที่เหลือคงไว้ตามเดิม เช่น "amanita farinosa" -> "Amanita farinosa")
"""

import json
import os
import re
import statistics
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from species_descriptions import SPECIES_DESCRIPTIONS

# ---------- ตั้งค่า path ----------
SOURCE_DIR = "source-data"
DATA_OUT_DIR = "data"
IMAGES_OUT_DIR = os.path.join("public", "images")

THAI_DATES = {
    1: "22 พฤศจิกายน 2568",
    2: "29 ธันวาคม 2568",
    3: "16 มกราคม 2569",
    4: "14 กุมภาพันธ์ 2569",
    5: "7 พฤษภาคม 2569",
    6: "18 มิถุนายน 2569",
    7: "18 กรกฎาคม 2569",
}

# คอลัมน์ (index เริ่มที่ 0)
COL_POINT = 0
COL_IMAGE = 1
COL_FAMILY = 2
COL_SCI_NAME = 3
COL_LOCAL_NAME = 4
COL_AMOUNT = 5
COL_GROUP = 6
COL_ORIGIN = 7
COL_ROLE = 8
COL_EDIBILITY = 9
COL_TEMP = 10        # K อุณหภูมิอากาศ
COL_HUMIDITY = 11    # L ความชื้นสัมพัทธ์อากาศ
COL_SOIL_PH = 12     # M pH ของดิน
COL_SOIL_TEMP = 13   # N อุณหภูมิในดิน
COL_SOIL_HUMIDITY = 14  # O ความชื้นของดิน

NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def normalize_name(name):
    if not name:
        return ""
    return re.sub(r"\s+", " ", str(name).strip())


def capitalize_first_letter(name):
    """ทำให้ตัวอักษรตัวแรกของชื่อวิทยาศาสตร์เป็นตัวพิมพ์ใหญ่ ส่วนที่เหลือคงเดิม
    เช่น 'amanita farinosa' -> 'Amanita farinosa'
         'Hexagonia tenuis (Fr.) Fr.' -> ไม่เปลี่ยน (ตัวแรกเป็นตัวใหญ่อยู่แล้ว)
    """
    if not name:
        return name
    if name[0].isalpha():
        return name[0].upper() + name[1:]
    return name


def find_data_end_row(ws):
    for row_idx in range(4, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if val and "วิเคราะห์ข้อมูลในภาพรวม" in str(val):
            return row_idx - 1
    return ws.max_row


def resolve_opc_path(base_dir, target):
    """แปลง Target ที่อ้างอิงใน .rels ให้เป็น path เต็มในไฟล์ zip
    รองรับทั้งแบบ relative (../drawings/x.xml) และแบบ absolute (/xl/drawings/x.xml)
    ซึ่งเป็นรูปแบบที่ openpyxl เขียนออกมาหลัง resave (ต่างจากไฟล์ต้นฉบับที่ยังไม่ผ่าน openpyxl)
    """
    if target.startswith("/"):
        return target.lstrip("/")
    # relative path: ไล่ระดับโฟลเดอร์ตาม base_dir
    parts = base_dir.split("/")
    for segment in target.split("/"):
        if segment == "..":
            if parts:
                parts.pop()
        elif segment == ".":
            continue
        else:
            parts.append(segment)
    return "/".join(parts)


def get_row_to_media_map(xlsx_path, sheet_index=1):
    with zipfile.ZipFile(xlsx_path) as z:
        sheet_rels_path = f"xl/worksheets/_rels/sheet{sheet_index}.xml.rels"
        if sheet_rels_path not in z.namelist():
            return {}
        with z.open(sheet_rels_path) as f:
            rels_xml = ET.parse(f)

        drawing_target = None
        for rel in rels_xml.findall(".//rel:Relationship", NS):
            if "drawing" in rel.attrib.get("Type", ""):
                drawing_target = rel.attrib["Target"]
        if not drawing_target:
            return {}

        drawing_path = resolve_opc_path("xl/worksheets", drawing_target)
        drawing_dir = "/".join(drawing_path.split("/")[:-1])
        drawing_rels_path = drawing_dir + "/_rels/" + drawing_path.split("/")[-1] + ".rels"

        rid_to_media = {}
        if drawing_rels_path in z.namelist():
            with z.open(drawing_rels_path) as f:
                drels_xml = ET.parse(f)
            for rel in drels_xml.findall(".//rel:Relationship", NS):
                target = rel.attrib["Target"]
                rid_to_media[rel.attrib["Id"]] = resolve_opc_path(drawing_dir, target)

        with z.open(drawing_path) as f:
            dxml = ET.parse(f)

        row_to_media = {}
        for anchor_tag in ["twoCellAnchor", "oneCellAnchor"]:
            for anchor in dxml.findall(f".//xdr:{anchor_tag}", NS):
                from_el = anchor.find("xdr:from", NS)
                if from_el is None:
                    continue
                row_el = from_el.find("xdr:row", NS)
                if row_el is None:
                    continue
                row = int(row_el.text) + 1

                blip = anchor.find(".//a:blip", NS)
                if blip is None:
                    continue
                embed_rid = blip.attrib.get(
                    "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"
                )
                media_path = rid_to_media.get(embed_rid)
                if media_path:
                    row_to_media.setdefault(row, []).append(media_path)
        return row_to_media


def safe_mean(values):
    nums = [v for v in values if isinstance(v, (int, float))]
    return round(statistics.mean(nums), 1) if nums else None


def process_round(round_num, source_path, images_out_root):
    wb = load_workbook(source_path)
    ws = wb.active

    end_row = find_data_end_row(ws)
    row_to_media = get_row_to_media_map(source_path, sheet_index=1)

    round_img_dir = os.path.join(images_out_root, f"round-{round_num}")
    os.makedirs(round_img_dir, exist_ok=True)

    with zipfile.ZipFile(source_path) as z:
        species = {}
        round_temps, round_humidities = [], []

        for row_idx in range(4, end_row + 1):
            point = ws.cell(row=row_idx, column=COL_POINT + 1).value
            sci_name_raw = ws.cell(row=row_idx, column=COL_SCI_NAME + 1).value
            if not point or not sci_name_raw:
                continue

            sci_key = normalize_name(sci_name_raw).lower()
            if not sci_key:
                continue

            air_temp = ws.cell(row=row_idx, column=COL_TEMP + 1).value
            air_humidity = ws.cell(row=row_idx, column=COL_HUMIDITY + 1).value
            soil_ph = ws.cell(row=row_idx, column=COL_SOIL_PH + 1).value
            soil_temp = ws.cell(row=row_idx, column=COL_SOIL_TEMP + 1).value
            soil_humidity = ws.cell(row=row_idx, column=COL_SOIL_HUMIDITY + 1).value

            if isinstance(air_temp, (int, float)):
                round_temps.append(air_temp)
            if isinstance(air_humidity, (int, float)):
                round_humidities.append(air_humidity)

            # ถือว่าเป็น "เห็ดดิน" ถ้าแถวนี้มีค่า M, N หรือ O อย่างน้อย 1 ค่า
            row_is_soil = any(
                isinstance(v, (int, float))
                for v in (soil_ph, soil_temp, soil_humidity)
            )

            row_images = []
            for i, media_path in enumerate(row_to_media.get(row_idx, [])):
                ext = os.path.splitext(media_path)[1] or ".jpg"
                filename = f"round-{round_num}-row-{row_idx:04d}-{i}{ext}"
                out_path = os.path.join(round_img_dir, filename)
                if not os.path.exists(out_path):
                    try:
                        with z.open(media_path) as src, open(out_path, "wb") as dst:
                            dst.write(src.read())
                    except KeyError:
                        continue
                row_images.append(f"/images/round-{round_num}/{filename}")

            if sci_key not in species:
                species[sci_key] = {
                    "scientificName": capitalize_first_letter(normalize_name(sci_name_raw)),
                    "localName": ws.cell(row=row_idx, column=COL_LOCAL_NAME + 1).value or "-",
                    "family": ws.cell(row=row_idx, column=COL_FAMILY + 1).value or "-",
                    "group": ws.cell(row=row_idx, column=COL_GROUP + 1).value or "-",
                    "habitat": ws.cell(row=row_idx, column=COL_ORIGIN + 1).value or "-",
                    "ecologicalRole": ws.cell(row=row_idx, column=COL_ROLE + 1).value or "-",
                    "edibility": ws.cell(row=row_idx, column=COL_EDIBILITY + 1).value or "ไม่มีข้อมูล",
                    "totalFound": 0,
                    "pointsFound": [],
                    "images": [],
                    "_airTemps": [],
                    "_airHumidities": [],
                    "_soilPHs": [],
                    "_soilTemps": [],
                    "_soilHumidities": [],
                    "_soilRowCount": 0,
                }

            rec = species[sci_key]
            rec["totalFound"] += 1
            if point not in rec["pointsFound"]:
                rec["pointsFound"].append(point)
            if row_images:
                rec["images"].extend(row_images)

            rec["_airTemps"].append(air_temp)
            rec["_airHumidities"].append(air_humidity)
            if row_is_soil:
                rec["_soilRowCount"] += 1
                rec["_soilPHs"].append(soil_ph)
                rec["_soilTemps"].append(soil_temp)
                rec["_soilHumidities"].append(soil_humidity)

    mushroom_list = []
    for rec in species.values():
        rec["pointsFoundCount"] = len(rec["pointsFound"])
        if not rec["images"]:
            rec["images"] = ["/images/placeholder-mushroom.png"]

        # ตัดสินว่าเป็นเห็ดดินหรือเห็ดตอไม้ จากสัดส่วนแถวที่มีข้อมูลดิน
        is_soil_mushroom = rec["_soilRowCount"] > 0
        rec["habitatType"] = "soil" if is_soil_mushroom else "wood"

        rec["airTemperature"] = safe_mean(rec["_airTemps"])
        rec["airHumidity"] = safe_mean(rec["_airHumidities"])

        if is_soil_mushroom:
            rec["soilPH"] = safe_mean(rec["_soilPHs"])
            rec["soilTemperature"] = safe_mean(rec["_soilTemps"])
            rec["soilHumidity"] = safe_mean(rec["_soilHumidities"])
        else:
            rec["soilPH"] = None
            rec["soilTemperature"] = None
            rec["soilHumidity"] = None

        # ลบคีย์ชั่วคราวที่ใช้คำนวณออกก่อน export
        for tmp_key in [
            "_airTemps", "_airHumidities", "_soilPHs",
            "_soilTemps", "_soilHumidities", "_soilRowCount",
        ]:
            rec.pop(tmp_key, None)

        # ผสมข้อมูล "ลักษณะทั่วไป" จากไฟล์ species_descriptions.py
        # จับคู่ด้วยชื่อวิทยาศาสตร์แบบ normalize (ตัดช่องว่าง + ตัวพิมพ์เล็ก)
        lookup_key = normalize_name(rec["scientificName"]).lower()
        rec["generalCharacteristics"] = SPECIES_DESCRIPTIONS.get(
            lookup_key, "ยังไม่มีข้อมูลลักษณะทั่วไปสำหรับเห็ดชนิดนี้"
        )

        mushroom_list.append(rec)

    mushroom_list.sort(key=lambda m: m["scientificName"].lower())

    avg_temp = safe_mean(round_temps)
    avg_humidity = safe_mean(round_humidities)

    return {
        "round": round_num,
        "date": THAI_DATES.get(round_num, ""),
        "speciesCount": len(mushroom_list),
        "mushrooms": mushroom_list,
        "avgTemperature": avg_temp,
        "avgHumidity": avg_humidity,
    }


def main():
    os.makedirs(DATA_OUT_DIR, exist_ok=True)
    os.makedirs(IMAGES_OUT_DIR, exist_ok=True)

    summary = []

    for round_num in range(1, 8):
        source_path = os.path.join(SOURCE_DIR, f"{round_num}.xlsx")
        if not os.path.exists(source_path):
            print(f"⚠️  ไม่พบไฟล์ {source_path} ข้ามรอบนี้ไป")
            continue

        print(f"กำลังประมวลผลรอบที่ {round_num} ...")
        result = process_round(round_num, source_path, IMAGES_OUT_DIR)

        out_json_path = os.path.join(DATA_OUT_DIR, f"round-{round_num}.json")
        with open(out_json_path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)

        with_image_count = sum(
            1 for m in result["mushrooms"]
            if m["images"] != ["/images/placeholder-mushroom.png"]
        )
        soil_count = sum(1 for m in result["mushrooms"] if m["habitatType"] == "soil")
        wood_count = sum(1 for m in result["mushrooms"] if m["habitatType"] == "wood")

        print(f"  -> พบ {result['speciesCount']} ชนิด "
              f"(มีรูปจริง {with_image_count}, เห็ดดิน {soil_count}, เห็ดตอไม้ {wood_count}), "
              f"เฉลี่ยอุณหภูมิ {result['avgTemperature']}°C, "
              f"เฉลี่ยความชื้น {result['avgHumidity']}%")

        summary.append({
            "round": round_num,
            "date": result["date"],
            "speciesCount": result["speciesCount"],
            "avgTemperature": result["avgTemperature"],
            "avgHumidity": result["avgHumidity"],
        })

    with open(os.path.join(DATA_OUT_DIR, "summary.json"), "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print("\nเสร็จสิ้น! ไฟล์ข้อมูลอยู่ในโฟลเดอร์ data/ และรูปอยู่ใน public/images/")


if __name__ == "__main__":
    main()